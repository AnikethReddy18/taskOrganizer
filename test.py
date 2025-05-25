from openpyxl import load_workbook

wb = load_workbook('/home/aniketh/Downloads/DSA.xlsx')
ws = wb.active

for row in ws.iter_rows():
    if row[1].value == None:
        break
    cell = row[1]
    print("Value:", cell.value)
    print("Hyperlink:", cell.hyperlink)
    print("Hyperlink target:", cell.hyperlink.target if cell.hyperlink else "No hyperlink")
    print("Formula:", cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else "Not a formula")
    print("===")
