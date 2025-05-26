from PyQt6.QtWidgets import (QApplication, QMainWindow, QFileDialog, QScrollArea,
                             QGridLayout, QLabel, QWidget, QPushButton)

import openpyxl
import random

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Window Management
        self.setWindowTitle("Task Organizer")
        self.resize(820, 400)

        # Initialize Rows
        self.rows =[]
        self.done_rows = []

        # Get Data
        self.open_file()
        self.get_data()

        # Display Data
        self.display_data()

    def open_file(self):
        # Open Excel File
        file_dialog = QFileDialog()
        self.file_path, _ = file_dialog.getOpenFileName(self, "Open Excel file", "", "Excel Files (*.xlsx)")
        self.workbook = openpyxl.load_workbook(self.file_path)
        sheet = self.workbook.active

        # Make "Done" Column if does not exist
        headers = [cell.value for cell in sheet[1]]
        if "Task Status" not in headers:
            sheet.cell(row=1, column=3).value = "Task Status"
            self.workbook.save(self.file_path)

    def get_data(self):
        # Make arrays for done rows and undone rows
        for i, row in enumerate(self.workbook.active.iter_rows()):
            if i == 0:
                continue
            if row[0].value == None:
                break
            title = row[0].value
            description = row[1].value
            link = row[1].hyperlink.target if row[1].hyperlink else None
            self.rows.append([title, description, link])
            if row[2].value == "Done":
                self.done_rows.append(1)
            else:
                self.done_rows.append(0)      
    
    def display_data(self):
        layout = QGridLayout()
        self.row_widgets = []

        # Labels for heading row
        for column_index,cell in enumerate(list(self.workbook.active.iter_rows())[0]):
                label = QLabel(cell.value)
                label.setWordWrap(True)
                layout.addWidget(label, 0, column_index)

        # Select Random row button        
        select_random_button = QPushButton("Select Random")
        select_random_button.pressed.connect(self.select_random_button_was_pressed)
        layout.addWidget(select_random_button, 0, 2, 1, 2)    
        
        # Add Undone rows to layout
        for row_index,row in enumerate(self.rows):
            # Skip display if task is done
            if self.done_rows[row_index] == 1:
                continue
            for column_index,cell in enumerate(row):
                label = QLabel(cell)
                label.setWordWrap(True)
                layout.addWidget(label, row_index+1, column_index)
            button = QPushButton("Done")
            button.clicked.connect(self.done_button_was_pressed)
            button.setProperty("index", row_index)
            self.columns = len(row)
            
            layout.addWidget(button, row_index+1, self.columns)
            self.row_widgets.append(button)

        # Make layout scrollable
        widget = QWidget()
        widget.setLayout(layout)
        self.scroll = QScrollArea()
        self.scroll.setWidget(widget)
        self.setCentralWidget(self.scroll)

    def done_button_was_pressed(self):
        # Get the index and edit the file to done
        index = self.sender().property('index')
        self.workbook.active.cell(row=index+2, column=self.columns).value = "Done"
        self.workbook.save(self.file_path)

        # Update done_rows
        self.done_rows[index] = 1
        # Update display to new data
        self.setCentralWidget(QWidget())
        self.display_data()
    
    def select_random_button_was_pressed(self):
        selected_row_index = random.randint(0, self.done_rows.count(0)-1)
        elem = self.row_widgets[selected_row_index]
        elem.setStyleSheet("color: red")
        self.scroll.ensureWidgetVisible(elem)

app = QApplication([])
window = MainWindow()
window.show()
app.exec()